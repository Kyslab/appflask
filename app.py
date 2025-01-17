# sửa sau khi copy
# app.config['MAIL_USERNAME'] = 'mergexceltool'
# app.config['MAIL_PASSWORD'] = 'uhuu dnwi kaiu days'
# app.config['MAIL_DEFAULT_SENDER'] = 'mergexceltool@gmail.com'

# khi dua lên heroku thì sửa thành Https:
# redirect_uri = url_for('authorize_google', _external=True, _scheme='https')

    # client_id='206611615101-g9kp571dagj69qn1b0ffb723c8qn9d7q.apps.googleusercontent.com',
    # client_secret='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',
import secrets
import os
from flask import Flask, jsonify, request, redirect, url_for, send_file, render_template, flash, session,send_from_directory
from werkzeug.utils import secure_filename
import openpyxl
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, BooleanField, SubmitField
from wtforms.validators import InputRequired, Length, Email, ValidationError,EqualTo
from flask_mail import Mail, Message
from random import randint
from urllib.parse import quote as url_quote
from authlib.integrations.flask_client import OAuth
from authlib.integrations.base_client.errors import OAuthError
from itsdangerous import URLSafeTimedSerializer, SignatureExpired
import paypalrestsdk
from decimal import Decimal, ROUND_UP
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import io
import dropbox
import pymysql
import random

app = Flask(__name__)
app.secret_key = 'supersecretkey'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['PROCESSED_FOLDER'] = 'processed'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx','xlsm'}

# @app.before_request
# def before_request():
#     if not request.is_secure and not app.debug:
#         url = request.url.replace("http://", "https://", 1)
#         return redirect(url, code=301)
    
# Database configuration
# app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://u7tbi7nfum7v0:pf5f201f6f9d9fbcc433d951c97a2556850b548a66deed2297b7a41bfe789d0b6@c5p86clmevrg5s.cluster-czrs8kj4isg7.us-east-1.rds.amazonaws.com:5432/d7ctfli13pgsc0'
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres.auigntrplygnjajymajy:Zoo8seekZoo8seek.&@aws-0-ap-southeast-1.pooler.supabase.com:6543/postgres'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
migrate = Migrate(app, db)

# Mail configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = 'mergexceltool'
app.config['MAIL_PASSWORD'] = 'uhuu dnwi kaiu days'
app.config['MAIL_DEFAULT_SENDER'] = 'mergexceltool@gmail.com'

mail = Mail(app)

# OAuth configuration
oauth = OAuth(app)
google = oauth.register(
    name='google',
    client_id='206611615101-g9kp571dagj69qn1b0ffb723c8qn9d7q.apps.googleusercontent.com',
    client_secret='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',
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    client_kwargs={
        'scope': 'openid email profile'
    }
)

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])
if not os.path.exists(app.config['PROCESSED_FOLDER']):
    os.makedirs(app.config['PROCESSED_FOLDER'])


# Cấu hình SDK của PayPal
paypalrestsdk.configure({
    "mode": "live",  # Chế độ sandbox để thử nghiệm, chuyển sang "live" để sản xuất
    "client_id": "AUxEs4nGbTtxSMnPjoqLABp2FgkZ54XzMNygJsOZZqqyqGjH7TLN3mkckVBl7l1QOAJbIlE-5c99XG0I",
    "client_secret": "EEpk9gH1CvZIEVc9EwKPcXghw1M61CfDoTz96zeIRbHK7j1B5U2ktIW-iXZnTIQHmpQ8iDyiCAzFnQ3B"
    # "mode": "sandbox",  # Chế độ sandbox để thử nghiệm, chuyển sang "live" để sản xuất
    # "client_id": "AQ4anPK0Dv374JlrBVbRHwm4JtPRJ6n1hk8myO9NbkXCLSu5bD3VjC4bDD-UOfYIAP0xHA1JbgN57vBA",
    # "client_secret": "ELPxD1aZbDCIub4HzdLxDUQ1nTLEQMBpNisFuW1MXo3kJ3VP9ZLDjL7lGoCMUWGNlxw8k9lAmBkYtAa3"

})
def roundup(number, decimals=0):
    factor = Decimal(10) ** -decimals
    return Decimal(number).quantize(factor, rounding=ROUND_UP)
# Define the User model
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(150), nullable=False)
    is_confirmed = db.Column(db.Boolean, default=False)
    balance = db.Column(db.Float, default=0.0)
    is_first_login = db.Column(db.Boolean, default=True)


class DepositForm(FlaskForm):
    amount = StringField('Amount', validators=[InputRequired(), Length(min=1, max=20)])

class RegisterForm(FlaskForm):
    email = StringField('Email', validators=[InputRequired(), Email(), Length(max=150)])
    password = PasswordField('Password', validators=[InputRequired(), Length(min=4, max=150)])
    
    def validate_email(self, email):
        user = User.query.filter_by(email=email.data).first()
        if user:
            raise ValidationError('Email is already taken.')

class LoginForm(FlaskForm):
    email = StringField('Email', validators=[InputRequired(), Email(), Length(max=150)])
    password = PasswordField('Password', validators=[InputRequired(), Length(min=4, max=150)])
    remember = BooleanField('Remember me')
    
class ConfirmEmailForm(FlaskForm):
    email = StringField('Email', validators=[InputRequired(), Email(), Length(max=150)])
    code = StringField('Confirmation Code', validators=[InputRequired(), Length(min=6, max=6)])

s = URLSafeTimedSerializer(app.config['SECRET_KEY'])
class RequestResetForm(FlaskForm):
    email = StringField('Email', validators=[InputRequired(), Email(), Length(max=150)])
    submit = SubmitField('Request Password Reset')

class ResetPasswordForm(FlaskForm):
    password = PasswordField('New Password', validators=[InputRequired(), Length(min=4, max=150)])
    confirm_password = PasswordField('Confirm Password', validators=[InputRequired(), EqualTo('password')])
    submit = SubmitField('Reset Password')

def send_reset_email(user):
    token = s.dumps(user.email, salt='password-reset-salt')
    msg = Message('Password Reset Request', sender='your-email@gmail.com', recipients=[user.email])
    link = url_for('reset_token', token=token, _external=True)
    msg.body = f'Please click the following link to reset your password: {link}'
    mail.send(msg)
# Google Sheets API setup
def get_sheet():
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_name('ggsheetapi.json', scope)
    client = gspread.authorize(creds)
    # Open your Google Sheet
    sheet = client.open('dulieu').sheet1  # sheet1 if you want to access the first sheet , không dùng được với file dulieu.xlsx
    return sheet
    # Mở bảng tính bằng ID hoặc URL
    # spreadsheet = client.open_by_key('1h_3FW5usYneH8J8_tfFdPSXrvmNeo4rcfdSkvy8sIus')
# Đường dẫn tới file credentials
GOOGLE_CREDENTIALS_FILE = 'ggsheetapi.json'
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']

# Tạo đối tượng credentials
creds = Credentials.from_service_account_file(GOOGLE_CREDENTIALS_FILE, scopes=SCOPES)
def get_file_id_by_name(file_name):
    # Xây dựng đối tượng dịch vụ Drive
    service = build('drive', 'v3', credentials=creds)
    
    # Tìm kiếm file theo tên
    query = f"name='{file_name}'"
    results = service.files().list(q=query, spaces='drive', fields="files(id, name)").execute()
    
    items = results.get('files', [])
    
    if not items:
        print(f"No file found with name {file_name}")
        return None
    else:
        # Giả sử tên file là duy nhất, ta lấy ID của file đầu tiên tìm thấy
        return items[0]['id']
def download_excel_file(file_id):
    # Xây dựng đối tượng dịch vụ Drive
    service = build('drive', 'v3', credentials=creds)
    
    # Lấy file từ Google Drive
    request = service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    
    done = False
    while not done:
        status, done = downloader.next_chunk()
        print(f"Download {int(status.progress() * 100)}%")
    
    fh.seek(0)
    return fh

# Đọc dữ liệu từ file Excel
def read_excel_data(file_stream):
    workbook = openpyxl.load_workbook(file_stream)
    sheet = workbook.active
    
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    print('data day :::')
    print(data)
    return data
# Thay thế bằng Access Token của bạn
DROPBOX_ACCESS_TOKEN = 'sl.B9TZafoPAqEWeeSNS9BJsz6qyAxqB15oGF4DOtIqgDiVn6M0n7Gj1PwLrK4veKPRPSP0tlPeaZwrAAWMX0hiXHz7Ny75QQyMDgC_VyO9FC69AtGzCpjIN5DV7HS5tUNZyS-DGZ373Sap'

# Hàm để tải file từ Dropbox và đọc dữ liệu
def download_excel_file_from_dropbox(file_path):
    dbx = dropbox.Dropbox(DROPBOX_ACCESS_TOKEN)
    
    try:
        # Tải file từ Dropbox về bộ nhớ tạm (in-memory)
        metadata, res = dbx.files_download(file_path)
        file_stream = io.BytesIO(res.content)
        
        # Đọc dữ liệu từ file Excel trong bộ nhớ
        workbook = openpyxl.load_workbook(file_stream)
        sheet = workbook.active
        
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        
        return data
    except dropbox.exceptions.ApiError as e:
        print(f"Failed to download file: {e}")
        return None
def connect_db():
    return pymysql.connect(
        host="localhost",
        user="root",
        password="",
        db="bpmsdb",
        cursorclass=pymysql.cursors.DictCursor
    )
@app.route('/phpindex', methods=['GET', 'POST'])
def phpindex():
    connection = connect_db()

    # Giả lập dữ liệu từ cơ sở dữ liệu
    # aboutus_description = "Chúng tôi cung cấp các dịch vụ chăm sóc sắc đẹp tốt nhất."
    # contact_address = "123 Đường ABC, Thành phố XYZ"
    # contact_phone = "123456789"
    # contact_email = "info@beautyparlour.com"

    if request.method == 'POST':
        print('da co post roii')
        name = request.form['name']
        email = request.form['email']
        services = request.form['services']
        adate = request.form['adate']
        atime = request.form['atime']
        phone = request.form['phone']
        aptnumber = random.randint(100000000, 999999999)

        try:
            with connection.cursor() as cursor:
                query = """
                INSERT INTO tblappointment (AptNumber, Name, Email, PhoneNumber, AptDate, AptTime, Services)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, (aptnumber, name, email, phone, adate, atime, services))
                connection.commit()

                # Get appointment number
                query = "SELECT AptNumber FROM tblappointment WHERE Email=%s AND PhoneNumber=%s"
                cursor.execute(query, (email, phone))
                result = cursor.fetchone()
                session['aptno'] = result['AptNumber']
                return redirect('/thank-you')
            print("ok")

        except Exception as e:
            flash('Something went wrong. Please try again.')
            print(e)

    # Fetch services for the dropdown
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM tblservices")
        services = cursor.fetchall()

        # Truy vấn về 'aboutus'
        cursor.execute("SELECT PageDescription FROM tblpage WHERE PageType='aboutus'")
        aboutus_result = cursor.fetchone()
        aboutus_description = aboutus_result['PageDescription'] if aboutus_result else ""

        # Truy vấn về 'contactus'
        cursor.execute("SELECT PageDescription, MobileNumber, Email FROM tblpage WHERE PageType='contactus'")
        contact_result = cursor.fetchone()
        if contact_result:
            contact_address = contact_result['PageDescription']
            contact_phone = contact_result['MobileNumber']
            contact_email = contact_result['Email']
        else:
            contact_address = ""
            contact_phone = ""
            contact_email = ""

    connection.close()
    return render_template('phpindex.html', services=services, aboutus_description=aboutus_description,
                           contact_address=contact_address,
                           contact_phone=contact_phone,
                           contact_email=contact_email)

@app.route('/thank-you')
def thank_you():
    return "Thank you for your appointment."

@app.route('/searchdropbox', methods=['GET', 'POST'])
def searchdropbox():
    if request.method == 'POST':
        search_id = request.form['id']
        file_path = "/dulieuTracuuweb.xlsx"  # Đường dẫn đến file Excel trên Dropbox
        
        # Tải file từ Dropbox
        data = download_excel_file_from_dropbox(file_path)
        
        if data:
            # Tìm các dòng có ID khớp với ID tìm kiếm
            results = [row for row in data if str(row[0]) == search_id]
            
            if not results:
                flash(f'No results found for ID {search_id}', 'warning')
            return render_template('resultsdropbox.html', results=results)
        else:
            flash('Failed to load data from Dropbox.', 'danger')
    
    return render_template('searchdropbox.html')
@app.route('/search2', methods=['GET'])
def search_page():
    return render_template('search2.html')
@app.route('/search2', methods=['POST']) # tải file excel trên drvie về và tra cứu ra nhiều KQ
def search2():
    
    search_id = request.form['id']
    # file_id = '1lao2NqrvD3to40wwxn3n-iznNyumwFKX'  # Thay thế bằng Google Drive file ID của bạn-đây là TK drive khác
    # file_id = '1-1G0AtX8Ejk0wwigPAUrsLKk8yf3SWWf'  # Thay thế bằng Google Drive file ID của bạn
    file_name = 'dulieu.xlsx'  # Thay thế bằng tên file bạn muốn tìm
        
        # Tìm file ID theo tên
    file_id = get_file_id_by_name(file_name)
    
    file_stream = download_excel_file(file_id)
    data = read_excel_data(file_stream)
    
    results = [row for row in data if str(row[0]) == search_id]
    
    if not results:
        return jsonify({'status': 'error', 'message': 'No results found for ID {}'.format(search_id)})

    return jsonify({'status': 'success', 'results': results}  )  
        
@app.route('/lookup', methods=['GET', 'POST']) #dùng để tra cứu google sheet cho 1 kết quả duy nhất
def lookup():
    sheet = get_sheet()
    result = None
    if request.method == 'POST':
        search_id = request.form['id']
        records = sheet.get_all_records()

        # Find the row with the matching id
        for row in records:
            if str(row['id']) == search_id:
                result = row
                break

    return render_template('lookup.html', result=result)
@app.route('/search', methods=['GET'])# GET hiển thị ra trang tra cứu googlesheet nhiều kết quả
def search_form():
    return render_template('search.html')
@app.route('/search', methods=['POST'])#thực hiện post tra cứu googlesheet nhiều kết quả
def search():
    data = request.json
    search_id = data.get('id')

    # Get all rows from the Google Sheet
    worksheet=get_sheet()
    rows = worksheet.get_all_records()
    # Find all rows with the matching ID
    matching_rows = [row for row in rows if str(row['id']) == str(search_id)]

    return jsonify(matching_rows)
@app.route('/deposit', methods=['GET', 'POST'])
def deposit():
    if 'user_id' not in session:
        flash('Please log in to access this page', 'warning')
        return redirect(url_for('login'))

    form = DepositForm()
    if form.validate_on_submit():
        amount = float(form.amount.data)
        user = User.query.get(session['user_id'])
        user.balance += amount
        db.session.commit()
        flash('Deposit successful. Your new balance is ${:.2f}'.format(user.balance), 'success')
        return redirect(url_for('account'))

    return render_template('deposit.html', form=form)

@app.route('/account')
def account():
    if 'user_id' not in session:
        flash('Please log in to access this page', 'warning')
        return redirect(url_for('login'))

    user = User.query.get(session['user_id'])
    return render_template('account.html', user=user)

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_request():
    if 'user_id' in session:
        return redirect(url_for('upload_and_list_files'))
    form = RequestResetForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user:
            send_reset_email(user)
        flash('An email has been sent with instructions to reset your password.', 'info')
        return redirect(url_for('reset_ok'))
    return render_template('reset_request.html', form=form)
@app.route('/reset_ok')
def reset_ok():
    return render_template('reset_ok.html')

@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_token(token):
    try:
        email = s.loads(token, salt='password-reset-salt', max_age=3600)
    except SignatureExpired:
        flash('The token is expired!', 'danger')
        return redirect(url_for('reset_request'))

    form = ResetPasswordForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=email).first()
        if user:
            user.password = form.password.data
            db.session.commit()
            flash('Your password has been updated! You can now log in.', 'success')
            return redirect(url_for('login'))
    return render_template('reset_token.html', form=form)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

@app.route('/register', methods=['GET', 'POST'])
def register():
    form = RegisterForm()
    if form.validate_on_submit():
        hashed_password = form.password.data
        new_user = User(email=form.email.data, password=hashed_password)
        db.session.add(new_user)
        db.session.commit()

        verification_code = randint(100000, 999999)
        session['email_verification'] = {'email': new_user.email, 'code': verification_code}

        msg = Message('Email Verification', sender='your-email@gmail.com', recipients=[new_user.email])
        msg.body = f'Your verification code is {verification_code}.'
        mail.send(msg)

        flash('Account created successfully. Please check your email for the verification code.', 'success')
        return redirect(url_for('confirm_email', email=new_user.email))
    return render_template('register.html', form=form)

@app.route('/confirm_email', methods=['GET', 'POST'])
def confirm_email():
    if 'email_verification' not in session:
        return redirect(url_for('register'))

    email = session['email_verification']['email']
    form = ConfirmEmailForm(email=email)

    if form.validate_on_submit():
        if form.code.data == str(session['email_verification']['code']):
            user = User.query.filter_by(email=email).first()
            user.is_confirmed = True
            db.session.commit()
            session.pop('email_verification', None)
            flash('Email confirmed successfully. Please log in.', 'success')
            return redirect(url_for('login'))
        else:
            flash('Invalid verification code. Please try again.', 'danger')
    return render_template('confirm_email.html', form=form, email=email)

@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()

        if len(form.password.data) < 4:
            flash('Password must be at least 4 characters long. Please check your password or log in with Google.', 'danger')
            return render_template('login.html', form=form)
              
        if user and user.password == form.password.data:
            session['user_id'] = user.id
            # flash('Logged in successfully', 'success')
            # Kiểm tra nếu đây là lần đăng nhập đầu tiên
            if user.is_first_login:
                user.balance += 3.0  # Cộng 3 USD vào tài khoản
                user.is_first_login = False  # Đánh dấu đã đăng nhập lần đầu
                db.session.commit()  # Lưu thay đổi vào cơ sở dữ liệu

                flash('Welcome! You have received a $3 bonus for your first login.', 'success')

            return redirect(url_for('upload_and_list_files'))
        else:
            flash('Login fail, try again or log in with Google.', 'danger')
    return render_template('login.html', form=form)

@app.route('/login/google')
def login_google():
    nonce = secrets.token_urlsafe()
    session['nonce'] = nonce
    redirect_uri = url_for('authorize_google', _external=True, _scheme='https')
    # redirect_uri = url_for('authorize_google', _external=True, _scheme='http')
    return google.authorize_redirect(redirect_uri, nonce=nonce)
@app.route('/authorize/google')
def authorize_google():
    try:
        token = google.authorize_access_token()
        nonce = session.pop('nonce', None)
        resp = google.parse_id_token(token, nonce=nonce)
        email = resp['email']

        user = User.query.filter_by(email=email).first()
        if not user:
            user = User(email=email, is_confirmed=True)
            db.session.add(user)
            db.session.commit()

        session['user_id'] = user.id
        # Kiểm tra nếu đây là lần đăng nhập đầu tiên
        if user.is_first_login:
            user.balance += 3.0  # Cộng 3 USD vào tài khoản
            user.is_first_login = False  # Đánh dấu đã đăng nhập lần đầu
            db.session.commit()  # Lưu thay đổi vào cơ sở dữ liệu

            flash('Welcome! You have received a $3 bonus for your first login.', 'success')

        flash('Logged in successfully with Google.', 'success')
        return redirect(url_for('upload_and_list_files'))
    except OAuthError as e:
        print(e.error)
        flash('Could not authorize with Google. Please try again.', 'danger')
        return redirect(url_for('login'))


@app.route('/logout')
def logout():
    session.pop('user_id', None)
    # flash('You have been logged out.', 'success')
    return redirect(url_for('login'))

@app.route('/mprivacy.html') # dùng đăng ksy google search
def mprivacy():
    return render_template('mprivacy.html')
@app.route('/google06a54390fa47b952.html') # dùng đăng ksy google search
def google06a54390fa47b952():
    return render_template('google06a54390fa47b952.html')
    # return redirect(url_for('google06a54390fa47b952'))
@app.route('/', methods=['GET', 'POST'])
def upload_and_list_files():
    if 'user_id' not in session:
        # flash('Please log in to access this page', 'warning')
        return redirect(url_for('login'))
    
    user = User.query.get(session['user_id'])
    user_folder = os.path.join(app.config['UPLOAD_FOLDER'], str(user.id))
    total_size = 0
    if not os.path.exists(user_folder):
        os.makedirs(user_folder)
    
    if request.method == 'POST':
        if 'files' in request.files:
            files = request.files.getlist('files')
            for file in files:
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    filepath = os.path.join(user_folder, filename)
                    file.save(filepath)
                    total_size += os.path.getsize(filepath)
            flash('Files uploaded successfully')
        elif 'merge' in request.form:
            remove_empty_rows = 'remove_empty' in request.form
            text_to_remove_list = request.form.getlist('text_to_remove')
            # Sử dụng list comprehension để loại bỏ các phần tử 'on' và ''
            text_to_remove_list = [text for text in text_to_remove_list if text != 'on' and text != '']

            merged_wb = openpyxl.Workbook()
            merged_ws = merged_wb.active
            merged_ws.title = 'Merged Data'

            file_names = request.form.getlist('file_names')

            total_size = sum(os.path.getsize(os.path.join(user_folder, f)) for f in file_names)
            total_cost = total_size * len(file_names) / 1024 / 1024 / 100

            if user.balance < total_cost:
                return jsonify({'flash_message': 'Insufficient balance. Please deposit more funds.', 'flash_category': 'danger','balance':user.balance,'total_cost':total_cost})
                # flash('Insufficient balance. Please deposit more money.', 'danger')
                
                # return redirect(url_for('upload_and_list_files'))

            user.balance -= total_cost
            db.session.commit()

            for index, filename in enumerate(file_names):
                selected_sheet = request.form.get(f'sheets_{index}')
                print(f'selected_sheet {selected_sheet}')
                
                if allowed_file(filename):
                    filepath = os.path.join(user_folder, filename)
                    wb = openpyxl.load_workbook(filepath, data_only=True)
                    # ws = wb.active
                    if selected_sheet in wb.sheetnames:
                        print('if selected_sheet in wb.sheetnames:')
                        ws = wb[selected_sheet]
                        print(f'ws:{ws}')
                        # ws = wb.active
                        for row in ws.iter_rows(values_only=True):
                            if remove_empty_rows and all(cell is None for cell in row):
                                continue
                            
                            if any(text in [str(cell) for cell in row if cell is not None] for text in text_to_remove_list):
                                continue
                        
                            merged_ws.append(row)

            merged_filename = 'merged_file.xlsx'
            merged_filepath = os.path.join(app.config['PROCESSED_FOLDER'], f"{user.id}_{merged_filename}")
            merged_wb.save(merged_filepath)
            # Gửi file qua email
            send_merged_file_via_email('doanvanky36k21@gmail.com', merged_filepath)

            flash('Files merged successfully')
            # return redirect(url_for('download_file', filename=f"{user.id}_{merged_filename}"))
            return jsonify({'balance': user.balance, 'message': 'Files merged successfully', 'file_url': url_for('download_file', filename=f"{user.id}_{merged_filename}")})
            

    files = os.listdir(user_folder)
    files = [f for f in files if allowed_file(f)]
    total_size = sum(os.path.getsize(os.path.join(user_folder, f)) for f in files)
    print(f'total_size {total_size}...len(files:{len(files)} ')
    # cost_per_file = 1  # Example cost per file
    # cost_per_mb = 0.1  # Example cost per MB
    # total_cost = (len(files) * cost_per_file) + ((total_size / (1024 * 1024)) * cost_per_mb)  # Calculate total cost
    total_cost =  total_size*len(files)/1024/1024/100


    file_sheets = {}
    for file in files:
        filepath = os.path.join(user_folder, file)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        file_sheets[file] = wb.sheetnames
    user.balance=roundup(user.balance,3)
    return render_template('upload_and_list.html', files=files, file_sheets=file_sheets, user=user,total_cost=roundup(total_cost,3))
def send_merged_file_via_email(recipient, filepath):
    msg = Message('Your Merged File', recipients=[recipient])
    with app.open_resource(filepath) as fp:
        msg.attach(os.path.basename(filepath), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", fp.read())
    mail.send(msg)
@app.route('/delete/<filename>', methods=['POST'])
def delete_file(filename):
    if 'user_id' not in session:
        flash('Please log in to access this page', 'warning')
        return redirect(url_for('login'))
    
    user = User.query.get(session['user_id'])

    file_path = os.path.join(app.config['UPLOAD_FOLDER'], str(user.id), filename)
    print(file_path)
    print('file_path')
    if os.path.exists(file_path):
        os.remove(file_path)
        flash(f'File {filename} deleted successfully')
    else:
        flash(f'File {filename} not found')
    # return jsonify({'flash_message': ' files deleted successfully'}), 200
    return redirect(url_for('upload_and_list_files'))

@app.route('/delete_all_files', methods=['POST'])
def delete_all_files():
    if 'user_id' not in session:
        return jsonify({'flash_message': 'Please log in to access this page', 'flash_category': 'warning'}), 401

    user = User.query.get(session['user_id'])
    user_folder = os.path.join(app.config['UPLOAD_FOLDER'], str(user.id))
    print(user_folder)
    print('user_folder')



    for filename in os.listdir(user_folder):
        file_path = os.path.join(user_folder, filename)
        if allowed_file(filename):
            os.remove(file_path)

    flash('All files deleted successfully', 'success')
    return redirect(url_for('upload_and_list_files'))
    # return jsonify({'flash_message': 'All files deleted successfully', 'flash_category': 'success'}), 200

@app.route('/download/<filename>')
def download_file(filename):
    if 'user_id' not in session:
        flash('Please log in to access this page', 'warning')
        return redirect(url_for('login'))

    file_path = os.path.join(app.config['PROCESSED_FOLDER'], filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return "File not found", 404
    
@app.route('/paypal-transaction-complete', methods=['POST'])
def paypal_transaction_complete():
    data = request.json
    order_id = data['orderID']
    details = data['details']
    amount = float(data['amount'])  # Get the amount from the request

    # Assuming you have user_id in session
    user_id = session['user_id']
    user = User.query.get(user_id)
    if user:
        # Update user balance with the specified amount
        user.balance += amount
        db.session.commit()
        flash('Your balance has been updated.', 'success')
    else:
        flash('User not found.', 'error')

    return jsonify({'status': 'success'}), 200
# @app.route('/paypal-transaction-complete', methods=['POST'])
# def paypal_transaction_complete():
#     data = request.json
#     order_id = data['orderID']
#     amount = float(data['amount'])  # Lấy số tiền từ request

#     # Xác thực giao dịch với PayPal
#     try:
#         payment = paypalrestsdk.Payment.find(order_id)
#         if payment.state == "approved":
#             # Assuming you have user_id in session
#             user_id = session['user_id']
#             user = User.query.get(user_id)

#             if user:
#                 # Cập nhật số dư của người dùng với số tiền đã nạp
#                 user.balance += amount
#                 db.session.commit()
#                 flash('Your balance has been updated.', 'success')
#             else:
#                 flash('User not found.', 'error')

#             return jsonify({'status': 'success'}), 200
#         else:
#             flash('Payment not approved.', 'error')
#             return jsonify({'status': 'failed', 'reason': 'Payment not approved'}), 400
#     except paypalrestsdk.ResourceNotFound as error:
#         flash('Payment not found.', 'error')
#         return jsonify({'status': 'failed', 'reason': 'Payment not found'}), 404


@app.route('/robots.txt')
def robots_txt():
    return send_from_directory(app.static_folder, 'robots.txt')

@app.route('/sitemap.xml')
def sitemap_xml():
    return send_from_directory(app.static_folder, 'sitemap.xml')



if __name__ == "__main__":
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
